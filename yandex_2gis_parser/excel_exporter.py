"""
Экспорт данных в Excel
"""
import os
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExcelExporter:
    """Класс для экспорта данных в Excel"""
    
    def __init__(self, config: dict):
        self.config = config
        self.output_path = config.get('output', {}).get('save_path', './output')
        self.filename_pattern = config.get('output', {}).get('filename_pattern', '{city}_{source}_{date}.xlsx')
        
        # Создаем директорию если не существует
        os.makedirs(self.output_path, exist_ok=True)
        
    def export_all(self, data: List[Dict], filename: str = None) -> str:
        """Экспорт всех данных в один файл Excel"""
        if not data:
            print("Нет данных для экспорта")
            return None
            
        if filename is None:
            date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"parsed_data_{date_str}.xlsx"
            
        filepath = os.path.join(self.output_path, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Данные"
        
        # Настройка стилей
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Определяем все уникальные ключи из данных
        all_keys = set()
        for item in data:
            all_keys.update(item.keys())
        
        # Порядок колонок
        column_order = [
            'city', 'search_query', 'source', 'name', 'category', 'address',
            'rating', 'reviews_count', 'phone', 'website', 'hours', 'url'
        ]
        
        # Добавляем недостающие ключи
        headers = []
        for key in column_order:
            if key in all_keys:
                headers.append(key)
                all_keys.remove(key)
        
        # Добавляем остальные ключи
        headers.extend(sorted(all_keys))
        
        # Заголовки
        header_names = {
            'city': 'Город',
            'search_query': 'Поисковый запрос',
            'source': 'Источник',
            'name': 'Название',
            'category': 'Категория',
            'address': 'Адрес',
            'rating': 'Рейтинг',
            'reviews_count': 'Отзывы',
            'phone': 'Телефон',
            'website': 'Сайт',
            'hours': 'Часы работы',
            'url': 'Ссылка'
        }
        
        # Вставляем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_names.get(header, header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
        # Данные
        for row_num, item in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = item.get(header, '')
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                
                # Выравнивание
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
            # Раскраска строк в зависимости от источника
            source = item.get('source', '')
            if source == 'yandex_maps':
                fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
            elif source == 'gis_2':
                fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).fill = fill
        
        # Автоширина колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
            
        # Фильтры
        ws.auto_filter.ref = ws.dimensions
        
        # Сохранение
        wb.save(filepath)
        print(f"Данные сохранены в файл: {filepath}")
        return filepath
        
    def export_by_city_source(self, all_data: Dict[str, List[Dict]]) -> List[str]:
        """Экспорт данных с разделением по городам и источникам"""
        saved_files = []
        
        for city, sources_data in all_data.items():
            for source, items in sources_data.items():
                if not items:
                    continue
                    
                date_str = datetime.now().strftime('%Y%m%d')
                filename = self.filename_pattern.format(
                    city=city.replace(' ', '_'),
                    source=source,
                    date=date_str
                )
                
                filepath = self.export_all(items, filename)
                if filepath:
                    saved_files.append(filepath)
                    
        return saved_files
